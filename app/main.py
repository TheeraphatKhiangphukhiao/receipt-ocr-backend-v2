from typing import Union
from fastapi import FastAPI, UploadFile, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse #สำหรับส่งไฟล์กลับไปเเบบ asynchronously
from openpyxl import Workbook #library ที่ออกเเบบมาเพื่อทำงานกับ excel โดยเฉพาะ โดยสามารถทำงานต่างๆที่เกี่ยวข้องกับ excel ได้เเทบทุกอย่าง
from openpyxl.utils import get_column_letter #สำหรับเเปลงตัวเลขที่เเทนตำเเหน่งคอลัมน์ให้กลายเป็นตัวอักษรที่ใช้เเทนคอลัมน์ใน excel
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side #สำหรับตกเเต่ง font ในตารางข้อมูล 
import pytesseract
import re
import pre_image #นำเข้าโมดูลจากไฟล์อื่นๆ



app = FastAPI()


origins = [ #รายชื่อเเหล่งที่มาที่ควรได้รับอนุญาตให้ทำการร้องขอข้ามเเหล่งที่มา
    "https://receipt-ocr-app-8c0a9.web.app",
    "http://localhost:5173",
    "http://202.28.34.197:8027"
]

app.add_middleware( #CORS or Cross-Origin Resource Sharing
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



@app.post("/write/excel")
async def write_excel(receipt_data: Request):

    file_path = "..\\write-file\\receipt.xlsx"

    try:
        
        receipt_data_json = await receipt_data.json() #สร้าง json

        columns_set = []
        cell_length = []

        for data_json in receipt_data_json['result']:
            for key in data_json:
                columns_set.append(key)
                cell_length.append(0)
            break

        columns_set = sorted(columns_set, key=lambda x: int(re.findall(r'\d+', x)[0]))
        print(columns_set)


        file = Workbook() #Workbook คือ object ที่หน้าตาเหมือนกับ Excel สร้างมาเพื่อเป็นตัวเเทนของไฟล์ Excel ที่เราต้องการ อ่าน-เขียน
        
        print(file)
        print(file.worksheets) #ถามถึงเเผ่นงาน
        print(file.sheetnames) #ถามชื่อเเผ่นงาน

        sheet = file['Sheet'] 

        custom_style = NamedStyle(name='custom style')
        
        custom_style.font = Font(name='TH SarabunPSK', size=14, bold=True) #กำหนด font
        custom_style.fill = PatternFill('solid', fgColor='FFFFCC')
        custom_style.border = Border(left=Side(border_style='thin', 
                                            color='B2B2B2'), 
                                            right=Side(border_style='thin', 
                                                        color='B2B2B2'), 
                                                        top=Side(border_style='thin', 
                                                                color='B2B2B2'),
                                                                bottom=Side(border_style='thin',
                                                                            color='B2B2B2')
                                    )
  
        i = 1
        for data_json in receipt_data_json['result']:
            j = 1
            for key in columns_set:
                sheet.cell(i, j).value = data_json[key]
                sheet.cell(i, j).style = custom_style

                j += 1
            i += 1

        for data_json in receipt_data_json['result']:
            i = 0
            for key in columns_set:
                if len(data_json[key]) > cell_length[i]:
                    cell_length[i] = len(data_json[key])
                i += 1
        
        print(cell_length) #เเสดงความยาวสูงสุดของเเต่ละ cell

        for i in range(len(columns_set)):
            columns_name = get_column_letter(i + 1) #เเปลงตัวเลขที่เเทนตำเเหน่งของคอลัมน์ไปเป็นตัวอักษรที่เเทนตำเเหน่งของคอลัมน์ที่ใช้ใน Excel โดยฟังก์ชันนี้จะรองรับหลายตัวอักษรด้วยเช่น AA BB
            sheet.column_dimensions[columns_name].width = (cell_length[i] + 5) #ขยายความกว้างของคอลัมน์
            
        
        file.save(file_path) #บันทึกออกมาเป็นไฟล์ Excel

        return FileResponse(file_path)
    
    except Exception:
        print('Error in this service')




@app.post("/image/tesseract")
async def extract_receipt_information(file: UploadFile):

    contents = await file.read() #อ่านข้อมูลทั้งหมดจากไฟล์
    img = pre_image.preprocessing(contents)


    result = [] #ประกาศตัวเเปรสำหรับเก็บข้อมูลของใบเสร็จตาม pattern ที่กำหนด
    current_index = 0
    company_name = ''

    bigc_pattern = r'^\d+.\d+\s+\d{13}'
    makro_pattern = r'^\d+\s+\d{13}'
    lotus_pattern = r'^\d{14}\s+\w+'

    try:
        text = pytesseract.image_to_string(img, lang='tha+eng', config='--psm 6') #เเปลงรูปภาพใบเสร็จไปเป็น text
        
        text_line = text.split('\n') #เเบ่งบรรทัดตามการขึ้นบรรทัดใหม่ \n

        for i in range(len(text_line)):

            if re.match(bigc_pattern, text_line[i]):
                print('Hello big c : ', text_line[i])

                company_name = 'บริษัท บิ๊กซี ซูเปอร์เซ็นเตอร์ จำกัด (มหาชน)'
                current_index = i

                break

            elif re.match(makro_pattern, text_line[i]):
                print('Hello makro : ', text_line[i])

                company_name = 'บริษัท ซีพี เเอ็กซ์ตร้า จำกัด (มหาชน)'
                current_index = i

                break

            elif re.match(lotus_pattern, text_line[i]):
                print('Hello lotus : ', text_line[i])

                company_name = 'บริษัท เอก-ชัย ดีสทริบิวชั่น ซิสเทม จำกัด'
                current_index = i

                break
            else:
                company_name = 'not found'

        if company_name != 'not found':

            #กำหนดชื่อบริษัท
            result.append({
                "item1": company_name
            })

            #กำหนด pattern สำหรับเก็บข้อมูล
            result.append({
                "item1": "จำนวน",
                "item2": "รหัสสินค้า",
                "item3": "รายการสินค้า",
                "item4": "หน่วยบรรจุ",
                "item5": "ราคาต่อหน่วย",
                "item6": "ส่วนลด บาท",
                "item7": "รหัสภาษี",
                "item8": "จำนวนเงิน (รวม VAT)"
            })

            print('ตำเเหน่งของ array ปัจจุบัน = ', current_index)
            print(text_line[current_index])

            result = await search_important_data(result, text_line, current_index, bigc_pattern, makro_pattern, lotus_pattern)
    
        else:
            print('รูปภาพไม่ถูกต้อง ต้องเป็น makro bigc lotus เท่านั้น')

            result = []

    except Exception:
        result = []
    
    
    return {
        "result": result
    }



async def search_important_data(result, text_line, current_index, bigc_pattern, makro_pattern, lotus_pattern):
    payment_amount = 0 #ประกาศตัวเเปรสำหรับเก็บ ยอดเงินชำระ
    count = 0

    i = current_index

    while i < len(text_line):

        number = 0 #ประกาศตัวเเปรสำหรับเก็บจำนวนของบรรทัดที่จะบวกไปหาข้อมูลที่เกินของรายการสินค้าในเเถวนั้นๆ
        over = "" #สำหรับเก็บข้อมูลที่เกินของรายการสินค้าในเเถวนั้นๆออกมา
        product_name = "" #ประกาศตัวเเปรสำหรับสร้างรายการสินค้าที่มีการ join ข้อมูลใน List

        if count >= 3:
            print('วนลูปจนถึงเเถวที่ไม่ต้องการ ทำการหยุดลูป โดยที่ count = ', count)
            break


        if re.match(bigc_pattern, text_line[i]):

            words = text_line[i].split() #เเบ่งข้อความตามการเว้นวรรค
            print(words)


            count2 = 1 #สำหรับนับจำนวนของบรรทัดที่จะบวกไปหาข้อมูลที่เกินของรายการสินค้าในเเถวนั้นๆ
            while True:
                if text_line[i+count2] == "": #ตรวจสอบว่าบรรทัดต่อๆไปของเเถวนั้นๆมีค่าว่างหรือไม่
                    count2 += 1 #ถ้ามีค่าว่างที่ให้ count2 บวกไปทีละ 1 เเล้ววนลูปไปเรื่อยๆจนกว่าจะเจอเเถวที่มีข้อมูล
                else:
                    number += count2 #ถ้าเจอเเถวมีข้อมูลก็ให้เก็บจำนวนของบรรทัดที่จะบวกไปหาข้อมูลที่เกินของรายการสินค้าในเเถวนั้นๆ เเละหยุดการทำงานของลูป
                    break


            if (not re.match(bigc_pattern, text_line[i+number])) and (not re.match(r'ออกแทนใบกํากับภาษีอย่างย่อ\s+เลขที่\s+\d+', text_line[i+number])):
                over = text_line[i+number] #นำข้อมูลที่เกินของรายการสินค้าในเเถวนั้นๆออกมา
                product_name = " ".join(words[2:-3]) + over #นำข้อมูลที่เกินมาต่อเข้ากับรายการสินค้าของตัวมัน

                i += number

            else:
                print("กรณีที่ไม่มีข้อความเกินจนขึ้นบรรทัดใหม่")
                product_name = " ".join(words[2:-3]) #เพิ่มรายการสินค้า, การ join หมายความว่านำข้อมูลใน List มารวมกันเเละเเทนที่ช่องที่ต่อกันด้วย " " หรือจะใส่ "-"


            result.append({
                "item1": words[0], #เพิ่มจำนวน
                "item2": words[1], #เพิ่มรหัสสินค้า
                "item3": product_name, #เพิ่มรายการสินค้า
                "item4": "", #เนื่องจากใบเสร็จ big c ไม่มี column สำหรับข้อมูลหน่วยบรรจุ ดังนั้นจึงใส่ค่าว่าง
                "item5": words[-3], #เพิ่มราคาต่อหน่วย (รวม VAT), การใช้ตัวเลขติดลบในการเข้าถึงสมาชิกของ List จะเป็นการเข้าถึงสมาชิกจากท้ายสุดมา -3 หมายถึงสมาชิกตัวที่สามจากด้านท้ายสุดของ List
                "item6": words[-2], #เพิ่มส่วนลด บาท, -2 หมายถึงสมาชิกตัวที่ 2 จากด้านท้ายสุดของ List
                "item7": "", #เนื่องจากใบเสร็จ big c ไม่มี column สำหรับข้อมูล VAT CODE ดังนั้นจึงใส่ค่าว่าง
                "item8": words[-1] #เพิ่มจำนวนเงิน (รวม VAT), -1 หมายถึงสมาชิกตัวเเรกจากด้านท้ายสุดของ List
            })

            payment_amount += float(words[-1]) #ทำการหาผลรวมสำหรับ ยอดเงินชำระ


        elif re.match(makro_pattern, text_line[i]):
            
            words = text_line[i].split() #เเบ่งข้อความตามการเว้นวรรค
            print(words)

            column4 = words[-4]
            column3 = words[-5]
            column4_data = ''
            column5_data = ''

            if column4.startswith(tuple("0123456789")) == True and column3.isnumeric() == False:
                column4_data = column4
                column5_data = " ".join(words[2:-4])

            elif column4.startswith(tuple("0123456789")) == False and column3.isnumeric() == True:
                column4_data = column3 + '' + column4
                column5_data = " ".join(words[2:-5])
            
            else:
                column4_data = column4
                column5_data = " ".join(words[2:-4])


            result.append({
                "item1": words[0], #เพิ่มจำนวน
                "item2": words[1], #เพิ่มรหัสสินค้า
                "item3": column5_data, #เพิ่มรายการสินค้า, การ join หมายความว่านำข้อมูลใน List มารวมกันเเละเเทนที่ช่องที่ต่อกันด้วย " " หรือจะใส่ "-"
                "item4": column4_data, #เพิ่มหน่วยบรรจุ, -4 หมายถึงสมาชิกตัวที่ 4 จากด้านท้ายสุดของ List
                "item5": words[-3], #เพิ่มราคาต่อหน่วย (รวม VAT), การใช้ตัวเลขติดลบในการเข้าถึงสมาชิกของ List จะเป็นการเข้าถึงสมาชิกจากท้ายสุดมา -3 หมายถึงสมาชิกตัวที่สามจากด้านท้ายสุดของ List
                "item6": "", #เนื่องจากใบเสร็จ makro ไม่มี column สำหรับข้อมูล ส่วนลด บาท ดังนั้นจึงใส่ค่าว่าง
                "item7": words[-2], #เพิ่มข้อมูล VAT CODE, -2 หมายถึงสมาชิกตัวที่สองจากด้านท้ายสุดของ List
                "item8": words[-1] #เพิ่มจำนวนเงิน (รวม VAT), -1 หมายถึงสมาชิกตัวเเรกจากด้านท้ายสุดของ List
            })

            payment_amount += float(words[-1]) #ทำการหาผลรวมสำหรับ ยอดเงินชำระ

        elif re.match(lotus_pattern, text_line[i]):
            
            words = text_line[i].split() #เเบ่งข้อความตามการเว้นวรรค
            print(words)

            result.append({
                "item1": words[-4], #เพิ่มจำนวน
                "item2": words[0], #เพิ่มรหัสสินค้า
                "item3": " ".join(words[1:-4]), #เพิ่มรายการสินค้า, การ join หมายความว่านำข้อมูลใน List มารวมกันเเละเเทนที่ช่องที่ต่อกันด้วย " " หรือจะใส่ "-"
                "item4": "", #เพิ่มหน่วยบรรจุ, -4 หมายถึงสมาชิกตัวที่ 4 จากด้านท้ายสุดของ List
                "item5": words[-3], #เพิ่มราคาต่อหน่วย (รวม VAT), การใช้ตัวเลขติดลบในการเข้าถึงสมาชิกของ List จะเป็นการเข้าถึงสมาชิกจากท้ายสุดมา -3 หมายถึงสมาชิกตัวที่สามจากด้านท้ายสุดของ List
                "item6": "", #เนื่องจากใบเสร็จ makro ไม่มี column สำหรับข้อมูล ส่วนลด บาท ดังนั้นจึงใส่ค่าว่าง
                "item7": "", #เพิ่มข้อมูล VAT CODE, -2 หมายถึงสมาชิกตัวที่สองจากด้านท้ายสุดของ List
                "item8": words[-2] + " บ" #เพิ่มจำนวนเงิน (รวม VAT), -1 หมายถึงสมาชิกตัวเเรกจากด้านท้ายสุดของ List
            })

            payment_amount += float(words[-2]) #ทำการหาผลรวมสำหรับ ยอดเงินชำระ

        else:
            count += 1


        i += 1

    #เพิ่มข้อมูล ยอดเงินชำระ
    result.append({
        "item1": "ยอดเงินชำระ",
        "item2": "",
        "item3": "",
        "item4": "",
        "item5": "",
        "item6": "",
        "item7": "",
        "item8": "{:.2f}".format(payment_amount)
    })

    return result